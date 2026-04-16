"""
CAP-01: Generacion de Archivos Excel (.xlsx)
=============================================

Validates that ILIAGPT.IO can generate production-quality Excel workbooks
with formulas, conditional formatting, charts, financial models, dashboards,
pivot tables, and data-cleaning pipelines.

Sub-capabilities under test:
  1.1  Formulas funcionales (VLOOKUP, SUMIF, IF, AVERAGE, PMT, etc.)
  1.2  Formato condicional con reglas de color
  1.3  Multiples hojas con datos aislados
  1.4  Graficos embebidos (bar, line, pie)
  1.5  Modelos financieros con analisis de escenarios
  1.6  Trackers de presupuesto con calculos automaticos
  1.7  Dashboards con KPIs y tendencias
  1.8  Tablas dinamicas (pivot simulation)
  1.9  Limpieza y transformacion de datos
"""
from __future__ import annotations

import pytest
from pathlib import Path
from openpyxl import load_workbook

from cowork_lib3 import (
    create_excel_with_formulas,
    create_excel_conditional_format,
    create_excel_multi_sheet,
    create_excel_with_chart,
    create_financial_model,
    create_budget_tracker,
    create_dashboard_xlsx,
    create_pivot_table_sim,
    clean_dataset,
)


@pytest.fixture
def out_dir(tmp_path):
    return tmp_path / "cap01_excel"


@pytest.fixture(autouse=True)
def _ensure_dir(out_dir):
    out_dir.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# 1.1 — Formulas funcionales
# ---------------------------------------------------------------------------

FORMULA_SETS = [
    ("SUM",         [["Item", "Q1", "Q2"], ["A", 10, 20], ["B", 30, 40]], {"D2": "=SUM(B2:C2)"},           "=SUM"),
    ("VLOOKUP",     [["Code", "Price"], ["A1", 100], ["A2", 200]],         {"D1": "=VLOOKUP(C1,A1:B3,2,FALSE)"}, "=VLOOKUP"),
    ("SUMIF",       [["Cat", "Val"], ["X", 10], ["Y", 20], ["X", 30]],    {"D1": '=SUMIF(A2:A4,"X",B2:B4)'},    "=SUMIF"),
    ("IF",          [["Score"], [85], [45]],                               {"B2": '=IF(A2>70,"Pass","Fail")'},    "=IF"),
    ("AVERAGE",     [["Val"], [10], [20], [30]],                           {"B1": "=AVERAGE(A1:A4)"},             "=AVERAGE"),
    ("COUNTIF",     [["Status"], ["Done"], ["Pending"], ["Done"]],         {"B1": '=COUNTIF(A2:A4,"Done")'},      "=COUNTIF"),
    ("INDEX_MATCH", [["ID", "Name"], [1, "Alice"], [2, "Bob"]],           {"D1": "=INDEX(B2:B3,MATCH(1,A2:A3,0))"}, "=INDEX"),
    ("PMT",         [["Principal", "Rate", "N"], [100000, 0.05, 360]],    {"D2": "=PMT(B2/12,C2,-A2)"},          "=PMT"),
    ("CONCATENATE", [["First", "Last"], ["John", "Doe"]],                 {"C2": '=CONCATENATE(A2," ",B2)'},     "=CONCATENATE"),
    ("MAX_MIN",     [["Val"], [5], [15], [10]],                            {"B1": "=MAX(A1:A4)", "C1": "=MIN(A1:A4)"}, "=MAX"),
]


@pytest.mark.file_generation
class TestExcelFormulas:
    """1.1 — Functional formula round-trip validation."""

    @pytest.mark.parametrize("desc,rows,formulas,prefix", FORMULA_SETS, ids=[f[0] for f in FORMULA_SETS])
    def test_formula_roundtrip(self, out_dir, desc, rows, formulas, prefix):
        """Formula string must survive write -> read cycle."""
        path = out_dir / f"formula_{desc}.xlsx"
        create_excel_with_formulas(path, rows, formulas)
        assert path.exists() and path.stat().st_size > 0, f"File not created for {desc}"
        wb = load_workbook(path)
        first_cell = list(formulas.keys())[0]
        val = str(wb.active[first_cell].value)
        assert val.startswith("="), f"Cell {first_cell} should contain formula, got: {val}"

    @pytest.mark.parametrize("n", range(1, 21))
    def test_multiple_formulas_per_sheet(self, out_dir, n):
        """A sheet should support N independent formula cells."""
        rows = [["x"]] + [[i] for i in range(1, 51)]
        formulas = {f"B{i+1}": f"=A{i+1}*2" for i in range(n)}
        path = out_dir / f"multi_formula_{n}.xlsx"
        create_excel_with_formulas(path, rows, formulas)
        wb = load_workbook(path)
        for cell_ref in formulas:
            assert str(wb.active[cell_ref].value).startswith("=")

    @pytest.mark.parametrize("i", range(20))
    def test_nested_if_formulas(self, out_dir, i):
        """Nested IF formulas must preserve full expression."""
        rows = [["Val"], [i * 10 + 5]]
        formulas = {"B2": f'=IF(A2>{i*5},"High",IF(A2>{i*2},"Med","Low"))'}
        path = out_dir / f"nested_{i}.xlsx"
        create_excel_with_formulas(path, rows, formulas)
        assert "=IF" in str(load_workbook(path).active["B2"].value)


# ---------------------------------------------------------------------------
# 1.2 — Formato condicional
# ---------------------------------------------------------------------------

@pytest.mark.file_generation
class TestExcelConditionalFormat:
    """1.2 — Conditional formatting rules applied to data ranges."""

    @pytest.mark.parametrize("n_rows", [5, 10, 25, 50, 100])
    def test_conditional_rules_created(self, out_dir, n_rows):
        """At least one conditional formatting rule group must exist after save/load."""
        data = [[i * 3.7 for _ in range(3)] for i in range(n_rows)]
        path = out_dir / f"cond_{n_rows}.xlsx"
        create_excel_conditional_format(path, data)
        rules = load_workbook(path).active.conditional_formatting
        assert len(list(rules)) >= 1, "Expected conditional formatting rule group"

    @pytest.mark.parametrize("high,low", [(90, 10), (80, 20), (70, 30), (60, 40), (50, 25)])
    def test_threshold_variants(self, out_dir, high, low):
        data = [[i] for i in range(0, 101, 5)]
        path = out_dir / f"cond_thresh_{high}_{low}.xlsx"
        create_excel_conditional_format(path, data, threshold_high=high, threshold_low=low)
        assert path.stat().st_size > 0

    @pytest.mark.parametrize("cols", range(1, 8))
    def test_multi_column_formatting(self, out_dir, cols):
        data = [[j * 10 + i for j in range(cols)] for i in range(20)]
        path = out_dir / f"cond_cols_{cols}.xlsx"
        create_excel_conditional_format(path, data)
        assert load_workbook(path).active.max_column == cols


# ---------------------------------------------------------------------------
# 1.3 — Multiples hojas
# ---------------------------------------------------------------------------

@pytest.mark.file_generation
class TestExcelMultiSheet:
    """1.3 — Multi-tab workbook creation with data isolation."""

    @pytest.mark.parametrize("n_sheets", range(2, 16))
    def test_sheet_count(self, out_dir, n_sheets):
        sheets = {f"Sheet_{i}": [["k", "v"], [f"row{i}", i]] for i in range(n_sheets)}
        path = out_dir / f"multi_{n_sheets}.xlsx"
        create_excel_multi_sheet(path, sheets)
        assert len(load_workbook(path).sheetnames) == n_sheets

    @pytest.mark.parametrize("i", range(10))
    def test_data_isolation_between_sheets(self, out_dir, i):
        """Data written to Sheet A must not leak to Sheet B."""
        sheets = {
            "Sales": [["Region", "Amount"], ["North", 100 + i]],
            "Costs": [["Category", "Amount"], ["Labor", 50 + i]],
        }
        path = out_dir / f"isolation_{i}.xlsx"
        create_excel_multi_sheet(path, sheets)
        wb = load_workbook(path)
        assert wb["Sales"].cell(2, 2).value == 100 + i
        assert wb["Costs"].cell(2, 2).value == 50 + i

    @pytest.mark.parametrize("i", range(10))
    def test_sheet_names_preserved(self, out_dir, i):
        names = [f"Data_{i}_{j}" for j in range(5)]
        sheets = {n: [["x", j]] for j, n in enumerate(names)}
        path = out_dir / f"names_{i}.xlsx"
        create_excel_multi_sheet(path, sheets)
        for n in names:
            assert n in load_workbook(path).sheetnames


# ---------------------------------------------------------------------------
# 1.4 — Graficos y visualizacion
# ---------------------------------------------------------------------------

@pytest.mark.file_generation
class TestExcelCharts:
    """1.4 — Embedded chart generation (bar, line, pie)."""

    @pytest.mark.parametrize("chart_type", ["bar", "line", "pie"])
    @pytest.mark.parametrize("n_cats", [3, 5, 8, 12])
    def test_chart_generation(self, out_dir, chart_type, n_cats):
        cats = [f"Cat_{i}" for i in range(n_cats)]
        series = {"Series_A": [i * 10 for i in range(n_cats)]}
        path = out_dir / f"chart_{chart_type}_{n_cats}.xlsx"
        create_excel_with_chart(path, cats, series, chart_type=chart_type)
        assert len(load_workbook(path).active._charts) >= 1, f"No chart found ({chart_type})"

    @pytest.mark.parametrize("n_series", [1, 2, 3, 5])
    def test_multi_series_chart(self, out_dir, n_series):
        cats = ["Q1", "Q2", "Q3", "Q4"]
        series = {f"Product_{i}": [i * 10 + j for j in range(4)] for i in range(n_series)}
        path = out_dir / f"multi_series_{n_series}.xlsx"
        create_excel_with_chart(path, cats, series, chart_type="bar")
        assert load_workbook(path).active.cell(1, 1).value == "Category"


# ---------------------------------------------------------------------------
# 1.5 — Modelos financieros
# ---------------------------------------------------------------------------

@pytest.mark.file_generation
class TestExcelFinancialModels:
    """1.5 — Multi-scenario financial models with P&L projections."""

    @pytest.mark.parametrize("base_rev", [100_000, 500_000, 1_000_000, 5_000_000])
    def test_model_structure(self, out_dir, base_rev):
        path = out_dir / f"finmodel_{base_rev}.xlsx"
        create_financial_model(path, base_rev, [0.05, 0.08, 0.10, 0.07, 0.06])
        wb = load_workbook(path)
        assert "FinancialModel" in wb.sheetnames
        assert "Scenarios" in wb.sheetnames
        assert wb["FinancialModel"].cell(1, 6).value == "Net Income"

    @pytest.mark.parametrize("cost_pct", [0.50, 0.60, 0.70, 0.80])
    def test_cost_sensitivity(self, out_dir, cost_pct):
        path = out_dir / f"fin_cost_{int(cost_pct*100)}.xlsx"
        create_financial_model(path, 1_000_000, [0.05] * 5, cost_pct=cost_pct)
        ws = load_workbook(path)["FinancialModel"]
        assert abs(ws.cell(2, 3).value - ws.cell(2, 2).value * cost_pct) < 1

    @pytest.mark.parametrize("tax", [0.10, 0.21, 0.30, 0.35])
    def test_tax_rate_applied(self, out_dir, tax):
        path = out_dir / f"fin_tax_{int(tax*100)}.xlsx"
        create_financial_model(path, 1_000_000, [0.05] * 3, tax_rate=tax)
        ws = load_workbook(path)["FinancialModel"]
        assert abs(ws.cell(2, 5).value - ws.cell(2, 4).value * tax) < 1

    @pytest.mark.parametrize("i", range(10))
    def test_three_scenarios_present(self, out_dir, i):
        path = out_dir / f"fin_scen_{i}.xlsx"
        create_financial_model(path, 500_000 + i * 100_000, [0.05 + i * 0.01] * 5)
        ws = load_workbook(path)["Scenarios"]
        scenarios = {ws.cell(r, 1).value for r in range(2, 5)}
        assert {"Bear", "Base", "Bull"} == scenarios


# ---------------------------------------------------------------------------
# 1.6 — Trackers de presupuesto
# ---------------------------------------------------------------------------

@pytest.mark.file_generation
class TestExcelBudgetTracker:
    """1.6 — Auto-calculated budget trackers with variance formulas."""

    @pytest.mark.parametrize("n_cats", [3, 5, 8, 12])
    def test_tracker_structure(self, out_dir, n_cats):
        cats = [f"Category_{i}" for i in range(n_cats)]
        path = out_dir / f"budget_{n_cats}.xlsx"
        create_budget_tracker(path, cats, [1000 + i * 500 for i in range(n_cats)], [900 + i * 450 for i in range(n_cats)])
        ws = load_workbook(path).active
        assert ws.cell(1, 4).value == "Variance"
        assert ws.cell(1, 6).value == "Status"

    @pytest.mark.parametrize("i", range(15))
    def test_variance_formulas(self, out_dir, i):
        path = out_dir / f"budget_formula_{i}.xlsx"
        create_budget_tracker(path, ["Mkt", "Eng", "Sales"], [5000, 8000, 3000], [4500, 9000, 2800])
        ws = load_workbook(path).active
        assert str(ws.cell(2, 4).value).startswith("="), "Variance must be a formula"
        assert str(ws.cell(2, 6).value).startswith("=IF"), "Status must be an IF formula"

    @pytest.mark.parametrize("i", range(10))
    def test_totals_row(self, out_dir, i):
        path = out_dir / f"budget_total_{i}.xlsx"
        create_budget_tracker(path, [f"D{j}" for j in range(4)], [1000]*4, [900+i*10]*4)
        ws = load_workbook(path).active
        assert ws.cell(6, 1).value == "TOTAL"
        assert str(ws.cell(6, 2).value).startswith("=SUM")


# ---------------------------------------------------------------------------
# 1.7 — Dashboards
# ---------------------------------------------------------------------------

@pytest.mark.file_generation
class TestExcelDashboards:
    """1.7 — Automated dashboards with KPIs and trend charts."""

    @pytest.mark.parametrize("n_kpis", [3, 5, 8, 10])
    def test_dashboard_sheets(self, out_dir, n_kpis):
        kpis = {f"KPI_{i}": 100 + i * 25 for i in range(n_kpis)}
        path = out_dir / f"dashboard_{n_kpis}.xlsx"
        create_dashboard_xlsx(path, kpis, {"Revenue": list(range(100, 1300, 100))})
        wb = load_workbook(path)
        assert "KPI_Dashboard" in wb.sheetnames and "Trends" in wb.sheetnames

    @pytest.mark.parametrize("i", range(10))
    def test_trend_chart_exists(self, out_dir, i):
        kpis = {"Revenue": 50_000 + i * 1000}
        path = out_dir / f"dash_trend_{i}.xlsx"
        create_dashboard_xlsx(path, kpis, {"Revenue": [j * 1000 for j in range(12)]})
        assert len(load_workbook(path)["Trends"]._charts) >= 1

    @pytest.mark.parametrize("i", range(10))
    def test_twelve_months_present(self, out_dir, i):
        path = out_dir / f"dash_months_{i}.xlsx"
        create_dashboard_xlsx(path, {"M": 100}, {"Sales": [j + i for j in range(12)]})
        assert load_workbook(path)["Trends"].max_row >= 13


# ---------------------------------------------------------------------------
# 1.8 — Tablas dinamicas (pivot)
# ---------------------------------------------------------------------------

@pytest.mark.file_generation
class TestExcelPivotTables:
    """1.8 — Simulated pivot table (cross-tab aggregation)."""

    @pytest.mark.parametrize("n_records", [10, 25, 50, 100])
    def test_pivot_sheets(self, out_dir, n_records):
        data = [{"region": f"R{i%3}", "product": f"P{i%4}", "sales": i * 100} for i in range(n_records)]
        path = out_dir / f"pivot_{n_records}.xlsx"
        create_pivot_table_sim(path, data, "region", "product", "sales")
        wb = load_workbook(path)
        assert "RawData" in wb.sheetnames and "Pivot" in wb.sheetnames

    @pytest.mark.parametrize("i", range(15))
    def test_pivot_totals_column(self, out_dir, i):
        data = [
            {"dept": "Sales", "qtr": "Q1", "amt": 100+i},
            {"dept": "Sales", "qtr": "Q2", "amt": 200+i},
            {"dept": "Eng",   "qtr": "Q1", "amt": 150+i},
            {"dept": "Eng",   "qtr": "Q2", "amt": 250+i},
        ]
        path = out_dir / f"pivot_total_{i}.xlsx"
        create_pivot_table_sim(path, data, "dept", "qtr", "amt")
        header = [load_workbook(path)["Pivot"].cell(1, c).value for c in range(1, 10) if load_workbook(path)["Pivot"].cell(1, c).value]
        assert "Total" in header

    @pytest.mark.parametrize("i", range(10))
    def test_raw_data_preserved(self, out_dir, i):
        data = [{"cat": f"C{j}", "sub": f"S{j%2}", "val": j * 10} for j in range(20)]
        path = out_dir / f"pivot_raw_{i}.xlsx"
        create_pivot_table_sim(path, data, "cat", "sub", "val")
        assert load_workbook(path)["RawData"].max_row == 21


# ---------------------------------------------------------------------------
# 1.9 — Limpieza y transformacion
# ---------------------------------------------------------------------------

@pytest.mark.data_science
class TestExcelDataCleaning:
    """1.9 — Data cleaning and transformation pipelines."""

    @pytest.mark.parametrize("i", range(20))
    def test_drop_nulls(self, i):
        data = [
            {"name": f"u_{i}", "email": f"u{i}@t.com"},
            {"name": "", "email": "bad@t.com"},
            {"name": f"u_{i+1}", "email": None},
            {"name": f"u_{i+2}", "email": f"u{i+2}@t.com"},
        ]
        assert len(clean_dataset(data, drop_nulls=True)) == 2

    @pytest.mark.parametrize("i", range(20))
    def test_deduplication(self, i):
        data = [{"id": "A", "v": 1+i}, {"id": "B", "v": 2+i}, {"id": "A", "v": 3+i}, {"id": "C", "v": 4+i}]
        cleaned = clean_dataset(data, dedup_key="id", drop_nulls=False)
        assert len(cleaned) == 3
        assert [r["id"] for r in cleaned] == ["A", "B", "C"]

    @pytest.mark.parametrize("i", range(15))
    def test_whitespace_stripping(self, i):
        data = [{"name": f"  user_{i}  ", "role": " admin "}]
        for r in clean_dataset(data, strip_whitespace=True, drop_nulls=False):
            assert r["name"] == r["name"].strip()

    @pytest.mark.parametrize("n", [10, 50, 100, 500])
    def test_large_dataset(self, n):
        data = [{"id": str(j % (n // 2)), "val": j, "tag": f" t{j} "} for j in range(n)]
        assert len(clean_dataset(data, dedup_key="id", strip_whitespace=True)) == n // 2

    @pytest.mark.parametrize("i", range(10))
    def test_combined_pipeline(self, i):
        data = [
            {"id": "1", "name": f"  Alice_{i} ", "email": f"a{i}@x.com"},
            {"id": "1", "name": f"Alice_{i}", "email": f"dup{i}@x.com"},
            {"id": "2", "name": "", "email": None},
            {"id": "3", "name": f" Bob_{i} ", "email": f"b{i}@x.com"},
        ]
        cleaned = clean_dataset(data, drop_nulls=True, dedup_key="id", strip_whitespace=True)
        assert len(cleaned) == 2
        assert cleaned[0]["name"] == f"Alice_{i}"
